# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl  import load_workbook
import datetime
import time
wb = Workbook()    #创建文件对象

# grab the active worksheet
ws = wb.active     #获取第一个sheet

# Data can be assigned directly to cells
ws['A1'] = 42      #写入数字
ws['B1'] = "你好"+"automation test" #写入中文（unicode中文也可）

# Rows can also be appended
ws.append([122, 12, 23])    #写入多个单元格

# Python types will automatically be converted

ws['A2'] = datetime.datetime.now()    #写入一个当前时间
#写入一个自定义的时间格式
ws['A3'] =time.strftime("%Y年%m月%d日 %H时%M分%S秒",time.localtime())

# Save the file
wb.save("sample.xlsx")

filename = "Test.xlsx"
wb = load_workbook(filename)

print(wb.sheetnames)#打印所有sheet

for sheet in wb:
	print(sheet.title)
	print(sheet.max_row,sheet.max_column)
	# for row in sheet.rows:
	sheet.insert_rows(1,1)
	# 	for cell in row:
	# 		print(cell.value)
	# 		cell.value = 33

wb.save(filename)