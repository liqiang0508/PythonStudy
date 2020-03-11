#!/usr/bin/python2
# -*- coding: UTF-8 -*-
import os
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
import json
import re
from lxml import etree
from openpyxl import Workbook
from openpyxl import load_workbook




def IsContainKey(dic,key):#成绩信息里面是否包含科目  
	for v in dic:
		v = v.decode("utf-8")
		# print type(v),type(key)
		if v==key:
			return True
	return False


def SaveScore(sheet,begin,end,row,cjDic):#查到每个人的成绩然后填上对应的成绩
	cell_range = sheet[begin:end]#科目名称的区间
	for cells in cell_range:
		for cell in cells:
			kcm = cell.value#科目名称
			kcm_col = cell.column#科目对应的列
			# print kcm
			if IsContainKey(cjDic,kcm):#成绩里面是否有成绩
				# print "have",kcm
				savecell = sheet.cell(row,kcm_col)
				savecell.value = cjDic[kcm.encode('utf-8')]#设置成绩
			else:#分数设置0
				savecell = sheet.cell(row,kcm_col)
				savecell.value = 0


def parseData(data,sfz,zkz,sheet,cur_row):
	html = etree.HTML(data)
	result = html.xpath('string(.)')
	try:
		jsondata = json.loads(result)
		rows = jsondata["rows"]
		return rows#返回成绩信息
	except Exception as e:
		return None
	

def getinfo(sfz,zkz,sheet,zybm,cur_row):#身份证，准考证,sheet,专业编码,当前行id
	# print("getinfo======",sfz)
	opt = webdriver.ChromeOptions()
	opt.headless = True
	driver = webdriver.Chrome(options=opt)
	# driver.implicitly_wait(5)
	driver.get(loginurl)
	
	time.sleep(3)
	usernum = driver.find_element_by_id("login")
	pwd = driver.find_element_by_id("password")
	loginbtn  = driver.find_element_by_id("btn_login")

# 登录
	pwdstr = sfz[-6:]#密码为身份证后6位
	usernum.send_keys(sfz)
	pwd.send_keys(pwdstr)
	loginbtn.send_keys(Keys.ENTER)
#
	time.sleep(1)
	zy_bm = zybm
	secondurl = "https://zk.sceea.cn/group/ks/-12?p_p_id=studentmain_WAR_signupportlet&p_p_lifecycle=1&p_p_state=normal&p_p_mode=view&p_p_col_id=column-1&p_p_col_count=1&_studentmain_WAR_signupportlet_javax.portlet.action=bytjListPage&ks_zkz="+str(zkz)+"&zy_bm="+str(zy_bm)+"&fx_bm=&page=1&rows=50"
	# print("secondurl==",secondurl)
	driver.get(secondurl)
	pageSource = driver.page_source
	
	time.sleep(0.5)
	driver.quit()
	return parseData(pageSource,sfz,zkz,sheet,cur_row)


def getSheetKms(sheet):#获取每个sheet的科目名称
	kms = {}
	for j in range(km_StartCol,km_EndCol+1):
		cell = sheet.cell(km_StartRow,j)
		column = cell.column
		kms[column] =  cell.value
		# print cell.value,column
	# print kms
	return kms #返回当前sheet的科目成绩的字典 列数:课程名




loginurl = "https://zk.sceea.cn/"
startrow = 4 #学生信息开始行数
km_StartCol = 12#科目开始列数
km_EndCol = 0#科目结束列数
km_StartRow = 3#科目开始行数





targeDir = os.getcwd()#目标文件夹
for dirpath,dirnames,filenames in os.walk(targeDir):#遍历目录下的所有文件
            for file in filenames:
                    if file.endswith("xlsx") and file!="Demon.xlsx":#是xlsx后缀的文件
						outputfile = file
						print "Start read***************"+outputfile
						wb = load_workbook(outputfile)
						sheetnames = wb.sheetnames
                    	
						for i in sheetnames:#遍历sheet
							print "Start Qurey sheet=======>"+i
							sheet = wb.get_sheet_by_name(i)
							km_EndCol = sheet.max_column#科目最大列数
							kms = getSheetKms(sheet)#课程信息
							# print len(kms),km_EndCol
							#获取一个sheet的课程名
							for row in sheet.iter_rows(min_row=startrow):
								
								cur_row = row[0].row#当前行
								studentName  = row[6].value#学生名称
								sfz = row[9].value#身份证
								zkz = row[8].value#准考证
								zymb = row[4].value#专业编码
								# print studentName, sfz,zkz,zymb,cur_row
								if zkz !=None:
									print studentName, sfz,zkz,zymb,cur_row
									data = getinfo(sfz,zkz,sheet,zymb,cur_row)#查成绩
									if data == None:#成绩为空火真没查到
										print studentName,"查询成绩失败"
										continue
									for k in kms:#表头里面的课程名称
										sheet.cell(cur_row,k).value = 0#先置0
										for j in data:#在判断获取的成绩里有没有对应的成绩
											kc_mc = j["kc_mc"]#课程名称
											cj = j["cj"]#课程名称

											if kms[k] == kc_mc:#课程名相等
												sheet.cell(cur_row,k).value = cj
												break
							print "End Qurey sheet=======>"+i
							
						print "end read***************"+outputfile
						wb.save(outputfile)
os.system("pause")